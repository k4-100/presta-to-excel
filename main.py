from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import mysql.connector
import os
from dotenv import load_dotenv

load_dotenv()



FILENAME = os.getenv('FILENAME')


def main():
    user = os.getenv('USER')
    password = os.getenv('PASSWORD')
    host = os.getenv("HOST")
    database = os.getenv("DATABASE")

    wb = Workbook()
    index = 1
    written_into_wb = False
    while True:
        query = input("następne zapytanie[j ENTER aby zakończyć]: ")
        # loop break
        if bool(query) == False:
            break

        if written_into_wb == False:
            written_into_wb = True


        print('###################################')
        body = []
        headers = []
        
        tabname = query.split("FROM")[1]
        if index > 1:
            ws = wb.create_sheet(f'{index} {tabname}')
        else:
            ws = wb.active
            ws.title = f'{index} {tabname}'

        print(f'wykonanywanie zapytania:\n{query} ')
        print("pobieranie danych z bazy...")

        cnx = mysql.connector.connect(
                user=user, 
                password=password,
                host=host,
                database=database
        )
        cursor = cnx.cursor()
        cursor.execute(query)
        
        for item in cursor:
            body.append( item )
        cursor.close()
        
        res = cnx._execute_query(query)
        for item in res["columns"]:
            headers.append(item[0])
        cnx.close()

        for x, item in enumerate(body):
            for y in range(0,len(item)):
                ws.cell( row = 1 + 1 + x, column = 1 + y, value = item[y])
        for x in range(0, len(headers)):
            ws.cell( row = 1, column = 1 + x, value = headers[x])
            ws.column_dimensions[get_column_letter(x+1)].width = 30

        index += 1 
        print("wkładanie danych do obiektu...\n")
        
        
    if written_into_wb:
        wb.save(f'{FILENAME}.xlsx')
        

if __name__ == '__main__':
    main()
else:
    print("Run from import")
